from datetime import date

from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from juntagrico.config import Config
from juntagrico.entity.member import Member
from juntagrico.entity.subs import Subscription
from juntagrico.entity.subtypes import SubscriptionType
from juntagrico.views.manage import MemberView
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from mehalsgmues.utils.utils import forum_notifications


@staff_member_required
def excel_export_subscriptions(request):
    filename = '{}_{}.xlsx'.format(Config.vocabulary('subscription_pl'), timezone.now().date())
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = Workbook()

    # Sheet 1: Subscriptions with prices
    ws1 = wb.active
    ws1.title = Config.vocabulary('subscription_pl')

    # header
    ws1.cell(1, 1, u"{}".format(Config.vocabulary('member_pl')))
    ws1.column_dimensions['A'].width = 40
    ws1.cell(1, 2, u"{}".format(_('E-Mail')))
    ws1.column_dimensions['B'].width = 30
    ws1.cell(1, 3, u"{}".format(_('Gesamtpreis [{}]').format(Config.currency())))
    ws1.column_dimensions['C'].width = 17
    for column, subs_type in enumerate(SubscriptionType.objects.all(), 4):
        ws1.cell(1, column, u"EAT {}".format(subs_type.price))
        ws1.column_dimensions[get_column_letter(column)].width = 17

    # data
    for row, subscription in enumerate(Subscription.objects.active(), 2):
        ws1.cell(row, 1, ", ".join([member.get_name() for member in subscription.current_members]))
        ws1.cell(row, 2, subscription.primary_member.email)
        ws1.cell(row, 3, subscription.price)
        for column, subs_type in enumerate(SubscriptionType.objects.all(), 4):
            ws1.cell(row, column, subscription.active_parts.filter(type__id=subs_type.id).count())

    ws1.freeze_panes = ws1['A2']
    wb.save(response)
    return response


@login_required
def ajax_notifications(request):
    notifications = forum_notifications(request.user)
    if notifications is None:
        notifications = u"\u00A0"
    return HttpResponse(notifications)


class MemberActiveView(MemberView):
    title = 'Alle Mitglieder'

    def get_queryset(self):
        self.queryset = Member.objects.exclude(share=None).filter(
            Q(share__termination_date__gte=date.today()) | Q(share__termination_date__isnull=True)
        ).distinct
        return super().get_queryset()
