import urllib
from datetime import timedelta
from itertools import accumulate

from dateutil.relativedelta import relativedelta
from dateutil.rrule import rrule, DAILY
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count, Avg, F
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.timezone import make_naive
from django.utils.translation import gettext_lazy as _
from juntagrico.config import Config
from juntagrico.dao.subscriptiondao import SubscriptionDao
from juntagrico.dao.subscriptiontypedao import SubscriptionTypeDao
from juntagrico.entity.jobs import ActivityArea
from juntagrico.entity.share import Share
from juntagrico.entity.subs import SubscriptionPart
from juntagrico.util.models import q_isactive
from juntagrico.util.temporal import start_of_business_year, end_of_business_year
from openpyxl import Workbook

from mehalsgmues.forms import DateRangeForm, CompareForm
from mehalsgmues.utils.stats import TemporalData, assignments_by, slots_by, assignments_by_subscription, members_with_assignments, get_assignment_progress
from mehalsgmues.utils.utils import date_from_get


@staff_member_required
def stats(request, trunc='week'):
    start_date = date_from_get(request, 'start_date', start_of_business_year())
    end_date = date_from_get(request, 'end_date', end_of_business_year())

    temporal_data = TemporalData(trunc, start_date, end_date)
    data = temporal_data.data_to_dict(assignments_by)
    done_jobs = list(data.values())
    available_slots = list(temporal_data.data_to_dict(slots_by, 'available').values())

    renderdict = {
        'trunc_name': temporal_data.trunc_adjective().capitalize() + 'e',
        'start_date': start_date,
        'end_date': end_date,
        'labels': [temporal_data.date_to_label(date) for date in data.keys()],
        'done_jobs': done_jobs,
        'available_slots': available_slots,
        'mobilization': [i / j if j > 0 else 0 for i, j in zip(done_jobs, available_slots)],
        'query': urllib.parse.urlencode(request.GET),
        'date_form': DateRangeForm(initial={'start_date': start_date, 'end_date': end_date})
    }
    return render(request, 'mag/stats.html', renderdict)


@staff_member_required
def stats_export(request):
    activity_area = ActivityArea.objects.filter(pk=request.GET.get('activity_area', None)).first()

    start_date = date_from_get(request, 'start_date', start_of_business_year())
    end_date = date_from_get(request, 'end_date', end_of_business_year())

    filename = '{}_{}_{}stats.xlsx'.format(
        start_date.strftime('%Y-%m-%d'),
        end_date.strftime('%Y-%m-%d'),
        str(activity_area.pk) + '_' if activity_area else ''
    )
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = Workbook()

    # Sheet 1: assignments by subscription
    ws1 = wb.active
    ws1.title = "assignments by subscription"

    # header
    ws1.cell(1, 1, u"{}".format(Config.vocabulary('member_pl')))
    ws1.column_dimensions['A'].width = 40
    ws1.cell(1, 2, u"{}".format(_('Arbeitseinsätze')))
    ws1.column_dimensions['B'].width = 17
    ws1.cell(1, 3, u"{}".format(_('{}-Grösse').format(Config.vocabulary('subscription'))))
    ws1.column_dimensions['C'].width = 17

    # data
    for row, subscription in enumerate(assignments_by_subscription(start_date, end_date, activity_area), 2):
        ws1.cell(row, 1, ", ".join([member.get_name() for member in subscription['subscription'].recipients]))
        ws1.cell(row, 2, subscription['assignments'])
        ws1.cell(row, 3, subscription['subscription'].totalsize)

    # Sheet 2: assignments per day
    ws2 = wb.create_sheet(title="assignments per day")

    # header
    ws2.cell(1, 1, u"{}".format(_('Datum')))
    ws2.column_dimensions['A'].width = 20
    ws2.cell(1, 2, u"{}".format(_('Arbeitseinsätze geleistet')))
    ws2.column_dimensions['B'].width = 17

    # data
    for row, assignment in enumerate(assignments_by('day', start_date, end_date, activity_area), 2):
        ws2.cell(row, 1, make_naive(assignment['day']))
        ws2.cell(row, 2, assignment['count'])

    # Sheet 3: slots by day
    ws3 = wb.create_sheet(title="slots per day")

    # header
    ws3.cell(1, 1, u"{}".format(_('Datum')))
    ws3.column_dimensions['A'].width = 20
    ws3.cell(1, 2, u"{}".format(_('Arbeitseinsätze ausgeschrieben')))
    ws3.column_dimensions['B'].width = 17

    # data
    for row, assignment in enumerate(slots_by('day', start_date, end_date, activity_area), 2):
        ws3.cell(row, 1, make_naive(assignment['day']))
        ws3.cell(row, 2, assignment['available'])

    # Sheet 4: assignments per member
    ws4 = wb.create_sheet(title="assignments per member")

    # header
    ws4.cell(1, 1, u"{}".format(Config.vocabulary('member')))
    ws4.column_dimensions['A'].width = 40
    ws4.cell(1, 2, u"{}".format(_('Arbeitseinsätze')))
    ws4.column_dimensions['B'].width = 17

    # data
    members = members_with_assignments(start_date, end_date, activity_area)
    for row, member in enumerate(members, 2):
        ws4.cell(row, 1, u"{}".format(member))
        ws4.cell(row, 2, member.assignments)

    wb.save(response)
    return response


@staff_member_required
def indexes(request):
    active_parts = SubscriptionPart.objects.filter(
        type__size__product__is_extra=False).filter(q_isactive()).filter(
        subscription__in=SubscriptionDao().all_active_subscritions()
    )
    types = SubscriptionTypeDao.get_all().filter(
        subscription_parts__in=active_parts
    ).annotate(num=Count('id')).order_by('-price')

    renderdict = dict(
        subscription_types=types,
        average_sub_price=active_parts.aggregate(avg=Avg('type__price'))['avg'],
        average_paid_sub_price=active_parts.filter(
            type__price__gt=0).aggregate(avg=Avg('type__price'))['avg'],
    )
    return render(request, 'mag/stats/indexes.html', renderdict)


@staff_member_required
def subscription_stats(request):
    start_date = date_from_get(request, 'start_date', start_of_business_year())
    end_date = date_from_get(request, 'end_date', end_of_business_year())

    labels = [start_date.strftime('%d.%m')]
    created = [SubscriptionPart.objects.filter(creation_date__lte=start_date).exclude(activation_date__lte=start_date).count()]
    active = [SubscriptionPart.objects.filter(activation_date__lte=start_date).exclude(cancellation_date__lte=start_date).count()]
    cancelled = [SubscriptionPart.objects.filter(cancellation_date__lte=start_date).exclude(deactivation_date__lte=start_date).count()]
    creations = dict(SubscriptionPart.objects.filter(creation_date__range=(start_date, end_date)).exclude(creation_date__gt=F('activation_date')).
                     values('creation_date').annotate(count=Count('id')).values_list('creation_date', 'count'))
    activations = dict(SubscriptionPart.objects.filter(activation_date__range=(start_date, end_date)).exclude(creation_date__gt=F('activation_date')).
                       values('activation_date').annotate(count=Count('id')).values_list('activation_date', 'count'))
    # because activation date may be set before creation, these are count separately: they count as +1 on actives but don't reduce created
    early_activations = dict(SubscriptionPart.objects.filter(creation_date__gt=F('activation_date'), activation_date__range=(start_date, end_date)).
                             values('activation_date').annotate(count=Count('id')).values_list('activation_date', 'count'))
    cancellations = dict(SubscriptionPart.objects.filter(cancellation_date__range=(start_date, end_date)).values('cancellation_date').annotate(count=Count('id')).values_list('cancellation_date', 'count'))
    deactivations = dict(SubscriptionPart.objects.filter(deactivation_date__range=(start_date, end_date)).values('deactivation_date').annotate(count=Count('id')).values_list('deactivation_date', 'count'))
    for day in rrule(DAILY, start_date + timedelta(1), until=end_date):
        day = day.date()
        labels.append(day.strftime('%d.%m'))
        a = activations.get(day, 0)
        c = cancellations.get(day, 0)
        created.append(creations.get(day, 0) - a)
        active.append(a + early_activations.get(day, 0) - c)
        cancelled.append(c - deactivations.get(day, 0))

    created = list(accumulate(created))
    active = list(accumulate(active))
    cancelled = list(accumulate(cancelled))

    renderdict = {
        'start_date': start_date,
        'end_date': end_date,
        'labels': labels,
        'active': active,
        'cancelled': cancelled,
        'created': created,
        'date_form': DateRangeForm(initial={'start_date': start_date, 'end_date': end_date})
    }
    return render(request, 'mag/stats/subscriptions.html', renderdict)


@staff_member_required
def shares(request):
    start_date = date_from_get(request, 'start_date', start_of_business_year())
    end_date = date_from_get(request, 'end_date', end_of_business_year())

    labels = [start_date.strftime('%d.%m')]
    created = [Share.objects.filter(creation_date__lte=start_date).exclude(paid_date__lte=start_date).count()]
    paid = [Share.objects.filter(paid_date__lte=start_date).exclude(cancelled_date=start_date).count()]
    cancelled = [Share.objects.filter(cancelled_date__lte=start_date).exclude(payback_date__lte=start_date).count()]
    creations = dict(Share.objects.filter(creation_date__range=(start_date, end_date)).exclude(creation_date__gt=F('paid_date')).
                     values('creation_date').annotate(count=Count('id')).values_list('creation_date', 'count'))
    payments = dict(Share.objects.filter(paid_date__range=(start_date, end_date)).exclude(creation_date__gt=F('paid_date')).
                    values('paid_date').annotate(count=Count('id')).values_list('paid_date', 'count'))
    # because paid date may be set before creation, these are count separately: they count as +1 on paid but don't reduce created
    early_payments = dict(Share.objects.filter(creation_date__gt=F('paid_date'), paid_date__range=(start_date, end_date)).
                             values('paid_date').annotate(count=Count('id')).values_list('paid_date', 'count'))
    cancellations = dict(Share.objects.filter(cancelled_date__range=(start_date, end_date)).values('cancelled_date').annotate(count=Count('id')).values_list('cancelled_date', 'count'))
    back_payments = dict(Share.objects.filter(payback_date__range=(start_date, end_date)).values('payback_date').annotate(count=Count('id')).values_list('payback_date', 'count'))
    for day in rrule(DAILY, start_date + timedelta(1), until=end_date):
        day = day.date()
        labels.append(day.strftime('%d.%m'))
        a = payments.get(day, 0)
        c = cancellations.get(day, 0)
        created.append(creations.get(day, 0) - a)
        paid.append(a + early_payments.get(day, 0) - c)
        cancelled.append(c - back_payments.get(day, 0))

    created = list(accumulate(created))
    paid = list(accumulate(paid))
    cancelled = list(accumulate(cancelled))

    renderdict = {
        'start_date': start_date,
        'end_date': end_date,
        'labels': labels,
        'paid': paid,
        'cancelled': cancelled,
        'created': created,
        'date_form': DateRangeForm(initial={'start_date': start_date, 'end_date': end_date})
    }
    return render(request, 'mag/stats/shares.html', renderdict)


@staff_member_required
def assignments(request):
    start_date = date_from_get(request, 'start_date', start_of_business_year())
    end_date = date_from_get(request, 'end_date', end_of_business_year())
    compare_years = int(request.GET.get('compare', 0))
    normalize = request.GET.get('normalize')

    data = []
    for i in range(0, compare_years + 1):
        data.insert(0, get_assignment_progress(start_date - relativedelta(years=i), end_date - relativedelta(years=i), normalize))

    return render(request, 'mag/stats/assignments.html', {
        'start_date': start_date,
        'end_date': end_date,
        'data': data,
        'date_form': DateRangeForm(initial={'start_date': start_date, 'end_date': end_date}),
        'compare_form': CompareForm(initial={'compare': compare_years, 'normalize': normalize})
    })
