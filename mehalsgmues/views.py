import urllib

import vobject
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Count, Avg
from django.shortcuts import render, redirect, get_object_or_404
from django.core.management import call_command
from django.urls import reverse
from django.utils.safestring import mark_safe
from django.utils.timezone import make_naive
from juntagrico.dao.subscriptiondao import SubscriptionDao
from juntagrico.dao.subscriptionproductdao import SubscriptionProductDao
from juntagrico.dao.subscriptiontypedao import SubscriptionTypeDao
from juntagrico.entity.depot import Depot
from juntagrico.entity.jobs import ActivityArea
from juntagrico.entity.share import Share
from juntagrico.entity.subs import Subscription, SubscriptionPart
from juntagrico.mailer import membernotification
from juntagrico.util import return_to_previous_location
from juntagrico.util.models import q_isactive
from juntagrico.util.views_admin import subscription_management_list
from openpyxl import Workbook

import base64
import hmac
import hashlib
from urllib import parse
from django.conf import settings


from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse

from juntagrico.models import Member

from juntagrico.dao.depotdao import DepotDao
from juntagrico.dao.listmessagedao import ListMessageDao
from juntagrico.util.temporal import weekdays, start_of_business_year, end_of_business_year
from juntagrico.config import Config
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from openpyxl.utils import get_column_letter

from mehalsgmues.forms import DateRangeForm
from mehalsgmues.utils.stats import assignments_by_subscription, \
    members_with_assignments, TemporalData, assignments_by, slots_by
from mehalsgmues.utils.utils import date_from_get, get_delivery_dates_of_month, forum_notifications


# API
@staff_member_required
def api_emaillist(request):
    """prints comma separated list of member emails"""
    sep = request.GET.get('sep', ', ')
    format = request.GET.get('format', 'plain')
    emails = sep.join(Member.objects.exclude(deactivation_date__lte=timezone.now().date()).values_list('email', flat=True))
    if format == 'plain':
        # just display for copy
        return HttpResponse(emails)
    else:
        # create file for download
        if format.find(';') == -1:
            content_type = f'text/{format}'
            file_type = format
        else:
            content_type, file_type = format.split(';')
        response = HttpResponse(content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="emails.{file_type}"'
        response.write(emails)
        return response


@staff_member_required
def api_vcf_contacts(request):
    members = Member.objects.exclude(deactivation_date__lte=timezone.now().date())
    cards = []
    for member in members:
        card = vobject.vCard()

        attr = card.add('n')
        attr.value = vobject.vcard.Name(family=member.last_name, given=member.first_name)

        attr = card.add('fn')
        attr.value = member.get_name()

        attr = card.add('email')
        attr.value = member.email

        if member.mobile_phone:
            attr = card.add('tel')
            attr.value = member.mobile_phone
            attr.type_param = 'cell'

        if member.phone:
            attr = card.add('tel')
            attr.value = member.phone
            attr.type_param = 'home'

        attr = card.add('org')
        attr.value = ["meh als gmües"]
        cards.append(card)

    response = HttpResponse(content_type='text/x-vcard')
    response['Content-Disposition'] = 'attachment; filename="members.vcf"'
    response.write("".join([card.serialize() for card in cards]))
    return response


@login_required
def nextcloud_profile(request):
    member = request.user.member
    membergroups = request.user.groups.values_list('name', flat=True)
    grouplist = list(membergroups)
    response = JsonResponse({'id': member.id,
                             'email': member.email,
                             'firstName': member.first_name,
                             'lastName': member.last_name,
                             'displayName': member.first_name + " " + member.last_name,
                             'roles': grouplist})
    return response


# pdf
def generate_pdf_dict():
    return {
        'subscriptions': SubscriptionDao.all_active_subscritions(),
        'products': SubscriptionProductDao.get_all(),
        'depots': DepotDao.all_depots_order_by_code(),
        'weekdays': {weekdays[weekday['weekday']]: weekday['weekday'] for weekday in
                     DepotDao.distinct_weekdays()},
        'messages': ListMessageDao.all_active()
    }


def other_recipients_names_w_linebreaks(self):
    members = self.recipients.exclude(email=self.primary_member.email)
    members = [str(member) for member in members]
    return mark_safe('<br>'.join([', '.join(members[x:x + 6]) for x in range(0, len(members), 6)]))


@staff_member_required
def list_mgmt(request, success=False):
    return render(request, 'mag/list_mgmt.html', {'success': success})


@staff_member_required
def list_generate(request, future=False):
    def delivery_dates(depot):
        return list(get_delivery_dates_of_month(depot.weekday - 1, int(request.GET.get('month', 0))))
    Depot.delivery_dates = delivery_dates
    call_command('generate_depot_list', force=True, future=future)
    return redirect(reverse('lists-mgmt-success'))


@staff_member_required
def stats(request, trunc='week'):
    start_date = date_from_get(request, 'start_date', start_of_business_year())
    end_date = date_from_get(request, 'end_date', end_of_business_year())

    temporal_data = TemporalData(trunc, start_date, end_date)
    data = temporal_data.data_to_dict(assignments_by)
    done_jobs = list(data.values())
    available_slots = list(temporal_data.data_to_dict(slots_by, 'available').values())

    renderdict = {
        'trunc_name': temporal_data.trunc_adjective().capitalize() + 'e',
        'start_date': start_date,
        'end_date': end_date,
        'labels': [temporal_data.date_to_label(date) for date in data.keys()],
        'done_jobs': done_jobs,
        'available_slots': available_slots,
        'mobilization': [i / j if j > 0 else 0 for i, j in zip(done_jobs, available_slots)],
        'query': urllib.parse.urlencode(request.GET),
        'date_form': DateRangeForm(initial={'start_date': start_date, 'end_date': end_date})
    }
    return render(request, 'mag/stats.html', renderdict)


@staff_member_required
def stats_export(request):
    activity_area = ActivityArea.objects.filter(pk=request.GET.get('activity_area', None)).first()

    start_date = date_from_get(request, 'start_date', start_of_business_year())
    end_date = date_from_get(request, 'end_date', end_of_business_year())

    filename = '{}_{}_{}stats.xlsx'.format(
        start_date.strftime('%Y-%m-%d'),
        end_date.strftime('%Y-%m-%d'),
        str(activity_area.pk) + '_' if activity_area else ''
    )
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = Workbook()

    # Sheet 1: assignments by subscription
    ws1 = wb.active
    ws1.title = "assignments by subscription"

    # header
    ws1.cell(1, 1, u"{}".format(Config.vocabulary('member_pl')))
    ws1.column_dimensions['A'].width = 40
    ws1.cell(1, 2, u"{}".format(_('Arbeitseinsätze')))
    ws1.column_dimensions['B'].width = 17
    ws1.cell(1, 3, u"{}".format(_('{}-Grösse').format(Config.vocabulary('subscription'))))
    ws1.column_dimensions['C'].width = 17

    # data
    for row, subscription in enumerate(assignments_by_subscription(start_date, end_date, activity_area), 2):
        ws1.cell(row, 1, ", ".join([member.get_name() for member in subscription['subscription'].recipients]))
        ws1.cell(row, 2, subscription['assignments'])
        ws1.cell(row, 3, subscription['subscription'].totalsize)

    # Sheet 2: assignments per day
    ws2 = wb.create_sheet(title="assignments per day")

    # header
    ws2.cell(1, 1, u"{}".format(_('Datum')))
    ws2.column_dimensions['A'].width = 20
    ws2.cell(1, 2, u"{}".format(_('Arbeitseinsätze geleistet')))
    ws2.column_dimensions['B'].width = 17

    # data
    for row, assignment in enumerate(assignments_by('day', start_date, end_date, activity_area), 2):
        ws2.cell(row, 1, make_naive(assignment['day']))
        ws2.cell(row, 2, assignment['count'])

    # Sheet 3: slots by day
    ws3 = wb.create_sheet(title="slots per day")

    # header
    ws3.cell(1, 1, u"{}".format(_('Datum')))
    ws3.column_dimensions['A'].width = 20
    ws3.cell(1, 2, u"{}".format(_('Arbeitseinsätze ausgeschrieben')))
    ws3.column_dimensions['B'].width = 17

    # data
    for row, assignment in enumerate(slots_by('day', start_date, end_date, activity_area), 2):
        ws3.cell(row, 1, make_naive(assignment['day']))
        ws3.cell(row, 2, assignment['available'])

    # Sheet 4: assignments per member
    ws4 = wb.create_sheet(title="assignments per member")

    # header
    ws4.cell(1, 1, u"{}".format(Config.vocabulary('member')))
    ws4.column_dimensions['A'].width = 40
    ws4.cell(1, 2, u"{}".format(_('Arbeitseinsätze')))
    ws4.column_dimensions['B'].width = 17

    # data
    members = members_with_assignments(start_date, end_date, activity_area)
    for row, member in enumerate(members, 2):
        ws4.cell(row, 1, u"{}".format(member))
        ws4.cell(row, 2, member.assignments)

    wb.save(response)
    return response


@staff_member_required
def indexes(request):
    active_parts = SubscriptionPart.objects.filter(
        type__size__product__is_extra=False).filter(q_isactive()).filter(
        subscription__in=SubscriptionDao().all_active_subscritions()
    )
    types = SubscriptionTypeDao.get_all().filter(
        subscription_parts__in=active_parts
    ).annotate(num=Count('id')).order_by('-price')

    renderdict = dict(
        subscription_types=types,
        average_sub_price=active_parts.aggregate(avg=Avg('type__price'))['avg'],
        average_paid_sub_price=active_parts.filter(
            type__price__gt=0).aggregate(avg=Avg('type__price'))['avg'],
    )
    return render(request, 'mag/stats/indexes.html', renderdict)


@staff_member_required
def excel_export_subscriptions(request):
    filename = '{}_{}.xlsx'.format(Config.vocabulary('subscription_pl'), timezone.now().date())
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = Workbook()

    # Sheet 1: Subscriptions with prices
    ws1 = wb.active
    ws1.title = Config.vocabulary('subscription_pl')

    # header
    ws1.cell(1, 1, u"{}".format(Config.vocabulary('member_pl')))
    ws1.column_dimensions['A'].width = 40
    ws1.cell(1, 2, u"{}".format(_('E-Mail')))
    ws1.column_dimensions['B'].width = 30
    ws1.cell(1, 3, u"{}".format(_('Gesamtpreis [{}]').format(Config.currency())))
    ws1.column_dimensions['C'].width = 17
    for column, subs_type in enumerate(SubscriptionTypeDao.get_all(), 4):
        ws1.cell(1, column, u"EAT {}".format(subs_type.price))
        ws1.column_dimensions[get_column_letter(column)].width = 17

    # data
    for row, subscription in enumerate(SubscriptionDao.all_active_subscritions(), 2):
        ws1.cell(row, 1, ", ".join([member.get_name() for member in subscription.recipients]))
        ws1.cell(row, 2, subscription.primary_member.email)
        ws1.cell(row, 3, subscription.price)
        for column, subs_type in enumerate(SubscriptionTypeDao.get_all(), 4):
            ws1.cell(row, column, subscription.active_parts.filter(type__id=subs_type.id).count())

    ws1.freeze_panes = ws1['A2']
    wb.save(response)
    return response


@login_required
def sso(request):
    payload = request.GET.get('sso')
    signature = request.GET.get('sig')

    if payload is None or signature is None:
        raise Exception('No SSO payload or signature. Please contact support if this problem persists.')

    # Validate the payload
    try:
        payload = bytes(parse.unquote(payload), encoding='utf-8')
        decoded = base64.decodebytes(payload).decode('utf-8')
        assert 'nonce' in decoded
        assert len(payload) > 0
    except AssertionError:
        raise Exception('Invalid payload. Please contact support if this problem persists.')

    key = bytes(settings.DISCOURSE_SSO_SECRET, encoding='utf-8')  # must not be unicode
    h = hmac.new(key, payload, digestmod=hashlib.sha256)
    this_signature = h.hexdigest()

    if not hmac.compare_digest(this_signature, signature):
        raise Exception('Invalid payload. Please contact support if this problem persists.')

    # Build the return payload
    qs = parse.parse_qs(decoded)
    params = {
        'nonce': qs['nonce'][0],
        'email': request.user.member.email,
        'external_id': request.user.id,
        'username': '%s.%s' % (request.user.member.first_name.lower(), request.user.member.last_name.lower()),
        'name': request.user.member.get_name(),
    }

    return_payload = base64.encodebytes(bytes(parse.urlencode(params), 'utf-8'))
    h = hmac.new(key, return_payload, digestmod=hashlib.sha256)
    query_string = parse.urlencode({'sso': return_payload, 'sig': h.hexdigest()})

    # Redirect back to Discourse
    url = '%s/session/sso_login' % settings.DISCOURSE_BASE_URL
    return HttpResponseRedirect('%s?%s' % (url, query_string))


@staff_member_required
def share_progress_preview(request):
    return render(request, 'mag/share_progress_preview.html')


@login_required
def bep(request):
    return render(request, 'mag/bep.html')


@login_required
def ajax_notifications(request):
    notifications = forum_notifications(request.user)
    if notifications is None:
        notifications = u"\u00A0"
    return HttpResponse(notifications)


@staff_member_required
def depot_changes(request):
    render_dict = {'change_date_disabled': True}
    return subscription_management_list(SubscriptionDao.subscritions_with_future_depots(), render_dict,
                                        'mag/management_lists/depot_changelist.html', request)


@staff_member_required
def depot_change_confirm(request, subscription_id):
    # Cheap copy of juntagrico util/subs.py activate_future_depots
    sub = get_object_or_404(Subscription, id=subscription_id)
    sub.depot = sub.future_depot
    sub.future_depot = None
    sub.save()
    emails = []
    for member in sub.recipients:
        emails.append(member.email)
    membernotification.depot_changed(emails, sub.depot)
    return return_to_previous_location(request)


@permission_required('juntagrico.view_share')
def share_unpaidlist(request):
    render_dict = {'change_date_disabled': True}
    return subscription_management_list(Share.objects.filter(paid_date__isnull=True).order_by('member'), render_dict,
                                        'mag/management_lists/share_unpaidlist.html', request)
